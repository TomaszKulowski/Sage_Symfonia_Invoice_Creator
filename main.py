"""This script is responsible for generating a document compatible with Sage Symfonia Handel
by combining data from an Excel report file, template, and additional information.

The script performs the following tasks:
1. Loads data from an Excel report file using the 'load_data' function.
2. Reads configuration settings from 'settings.json'.
3. Retrieves a document template from 'templates/symfonia.txt'.
4. Combines the template, data, and additional information to create a document.
5. Saves the resulting document to 'output.txt'.
"""
from json import load
from openpyxl import load_workbook
from typing import Generator

from device import RentedDevice
from templates.template import get_contractor, get_all_contractor_data, get_document_date


def load_data(report_filename: str) -> Generator:
    """Load data from an Excel report file and yield RentedDevice objects.

    Args:
        report_filename (str): The path to the Excel report file.

    Yields:
        RentedDevice: A generator that yields RentedDevice objects.
    """
    column_names = ['basic_fee', 'basic_black_copies', 'basic_color_copies', 'previous_black', 'previous_color',
                    'current_black', 'current_color', 'sn']
    workbook = load_workbook(filename=report_filename)
    sheet = workbook.worksheets[0]
    rows = sheet.max_row

    for row in range(2, rows + 1):
        rented_device = RentedDevice
        for column_name, column in zip(column_names, range(4, 12)):
            setattr(rented_device, column_name, sheet.cell(row=row, column=column).value)
        yield rented_device


def create_document():
    """Create an 'output.txt' file contains data to import in Sage Symfonia Handel.

    This function reads settings from 'settings.json', loads a template from 'templates/symfonia.txt',
    retrieves data using the 'load_data' function, and generates a document by combining the template,
    data, and additional information. The resulting document is saved to 'output.txt'.

    Note:
        - 'settings.json' must contain the necessary configuration information.
        - 'templates/symfonia.txt' should be a template for the document.

    Raises:
        FileNotFoundError: If 'settings.json' or 'templates/symfonia.txt' is not found.
    """
    with open('settings.json', encoding='utf-8') as file:
        settings = load(file)

    with open('templates/symfonia.txt') as file:
        content = file.read()[:-1]

    data = load_data(settings['report_filename'])
    content += get_document_date(settings['due_days'])
    for device in data:
        consumption = device().check_consumption()
        for position in consumption:
            content += position
    content += get_all_contractor_data(
        name=settings['contractor_name'],
        address=settings['address'],
        number=settings['number'],
        local=settings['local'],
        city=settings['city'],
        postal_code=settings['postal_code'],
        tax_number=settings['tax_number'],
    )
    content += '}\n'
    content += get_contractor(settings['contractor_name'])

    with open('output.txt', 'w', encoding='utf-8') as file:
        file.write(content)


if __name__ == '__main__':
    create_document()
